import os
import openpyxl
import matplotlib.pyplot as plt
import logging
from openpyxl.utils.exceptions import InvalidFileException

# Constants
STATS_FOLDER = 'stats'
SEMESTER_FILE = 'semestre.xlsx'
NOTES_FILE = 'notes_RT.xlsx'

# Logger setup
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Function to check and change the working directory
def check_and_change_directory():
    try:
        if not os.path.exists(STATS_FOLDER):
            logging.error(f"The '{STATS_FOLDER}' folder does not exist.")
            exit()
        os.chdir(STATS_FOLDER)
    except Exception as e:
        logging.error(f"Error accessing the '{STATS_FOLDER}' folder: {e}")
        exit()

# Function to validate the existence of necessary files
def validate_files():
    try:
        if not os.path.exists(SEMESTER_FILE):
            logging.error(f"The file '{SEMESTER_FILE}' is not in the '{STATS_FOLDER}' folder.")
            exit()

        if not os.path.exists(NOTES_FILE):
            logging.warning(f"The file '{NOTES_FILE}' is missing. Creating it now...")
            create_notes_file()
        else:
            update_notes_file()
    except Exception as e:
        logging.error(f"Error validating files: {e}")
        exit()

# Function to load Excel workbooks
def load_workbook(file_name, read_only=True):
    try:
        return openpyxl.load_workbook(file_name, data_only=True, read_only=read_only)
    except InvalidFileException:
        logging.error(f"Error: '{file_name}' is corrupted or invalid.")
        exit()
    except Exception as e:
        logging.error(f"Error loading the file '{file_name}': {e}")
        exit()

# Function to create the notes file
def create_notes_file():
    try:
        wb_semestre = load_workbook(SEMESTER_FILE)
        wb_notes = openpyxl.Workbook()

        for sheet_name in wb_semestre.sheetnames:
            sheet = wb_semestre[sheet_name]
            ws_notes = wb_notes.create_sheet(title=sheet_name)
            ws_notes.append(["Title", "Note1"])

            for row in range(3, sheet.max_row + 1):
                title = sheet.cell(row=row, column=1).value
                if title and isinstance(title, str):
                    ws_notes.append([title])

        if 'Sheet' in wb_notes.sheetnames:
            del wb_notes['Sheet']

        wb_notes.save(NOTES_FILE)
        wb_notes.close()
        wb_semestre.close()
        logging.info(f"The file '{NOTES_FILE}' was successfully created.")
    except Exception as e:
        logging.error(f"Error creating '{NOTES_FILE}': {e}")
        exit()

# Function to update the notes file
def update_notes_file():
    try:
        wb_semestre = load_workbook(SEMESTER_FILE)
        wb_notes = load_workbook(NOTES_FILE, read_only=False)

        for sheet_name in wb_semestre.sheetnames:
            sheet = wb_semestre[sheet_name]
            if sheet_name not in wb_notes.sheetnames:
                ws_notes = wb_notes.create_sheet(title=sheet_name)
                ws_notes.append(["Title", "Note1", "Note2", "Note3"])
            else:
                ws_notes = wb_notes[sheet_name]

            existing_titles = {ws_notes.cell(row=row, column=1).value for row in range(2, ws_notes.max_row + 1)}

            for row in range(3, sheet.max_row + 1):
                title = sheet.cell(row=row, column=1).value
                if title and isinstance(title, str) and title not in existing_titles:
                    ws_notes.append([title])

        wb_notes.save(NOTES_FILE)
        wb_notes.close()
        wb_semestre.close()
        logging.info(f"The file '{NOTES_FILE}' was successfully updated.")
    except Exception as e:
        logging.error(f"Error updating '{NOTES_FILE}': {e}")
        exit()

# Function to validate data consistency
def validate_data(sheet):
    if sheet.max_row < 3:
        logging.warning(f"Sheet '{sheet.title}' has insufficient rows for processing.")
        return False
    if sheet.max_column < 2:
        logging.warning(f"Sheet '{sheet.title}' has insufficient columns for processing.")
        return False
    return True

# Function to calculate averages per UE
def calculate_averages(wb_file, wb_note):
    try:
        UE_S = []
        notes_dict = {}

        for sheet_name in wb_note.sheetnames:
            sheet_notes = wb_note[sheet_name]

            if not validate_data(sheet_notes):
                continue

            for row in range(2, sheet_notes.max_row + 1):
                title = str(sheet_notes.cell(row=row, column=1).value)
                notes_dict[title] = []
                for col in range(2, sheet_notes.max_column + 1):
                    note = sheet_notes.cell(row=row, column=col).value
                    if isinstance(note, (int, float)):
                        notes_dict[title].append(note)

        for sheet_name in wb_file.sheetnames:
            sheet_file = wb_file[sheet_name]
            if not validate_data(sheet_file):
                continue

            for col in range(1, sheet_file.max_column + 1):
                for row in range(1, sheet_file.max_row + 1):
                    cell = sheet_file.cell(row=row, column=col)
                    if cell.value and cell.data_type == 's' and cell.value.startswith('UE'):
                        UE = cell.value
                        UEden, UEnom = 0, 0

                        for i in range(3, sheet_file.max_row + 1):
                            coefficient = sheet_file.cell(row=i, column=col).value
                            title = str(sheet_file.cell(row=i, column=1).value)
                            if coefficient and title and title in notes_dict:
                                for note in notes_dict[title]:
                                    UEnom += note * coefficient
                                    UEden += coefficient

                        if UEden != 0:
                            UE_S.append((UE, UEnom / UEden))

        return UE_S
    except Exception as e:
        logging.error(f"Error calculating averages: {e}")
        return []

# Function to plot results
def plot_results(values):
    try:
        plt.figure()
        plt.title("Results of BUT RT")
        plt.ylim(0, 20)
        plt.axhline(10, color="Red", linestyle='--', label='Pass Threshold')
        plt.axhline(8, color="Orange", linestyle='--', label='Retake Threshold')

        for UE, avg in values:
            color = "Green" if avg >= 10 else "Orange" if avg >= 8 else "Red"
            plt.bar(UE, avg, color=color)

        plt.legend()
        plt.show()
    except Exception as e:
        logging.error(f"Error generating the graph: {e}")

# Main function
def main():
    logging.info(f"Current directory: {os.getcwd()}")
    check_and_change_directory()
    validate_files()

    wb_file = load_workbook(SEMESTER_FILE)
    wb_note = load_workbook(NOTES_FILE)

    UE_averages = calculate_averages(wb_file, wb_note)
    if UE_averages:
        plot_results(UE_averages)
    else:
        logging.warning("No averages calculated. Check your files.")

    wb_file.close()
    wb_note.close()

if __name__ == "__main__":
    main()